from unidecode import unidecode
from .listas import lista_consolidado,meses
from .dbs import *
from .sql_sicad import sicad, sicad_desaparecimento
from .sql_sisp import sisp
import flet as ft
import psycopg2
import cx_Oracle
import openpyxl
import re


class qauntum_page(ft.View):
    def __init__(self, page: ft.Page):
        super().__init__(route="/qauntum",padding=40)




        file_picker = ft.FilePicker()
        page.overlay.append(file_picker)



        sugestoes = ft.ListView(
            visible=False, 
            width=345, 
            height=100,
            )




        def aviso(message, color=""):
            if color == "red":
                alerta = ft.AlertDialog(
                    title=ft.Text(
                        "Erro", 
                        color='#FFFFFF',
                        font_family='inter'),
                    content=ft.Text(
                        message,
                        size=15,
                        font_family='inter',
                        color='#FFFFFF',),
                    actions=[ft.TextButton(
                        "OK", 
                        on_click=lambda e: page.close(alerta),
                        style=ft.ButtonStyle(
                            color='#0064FF',
                            bgcolor={
                                    ft.ControlState.DEFAULT: "#FFFFFF",
                                    ft.ControlState.FOCUSED: "#5caee0",
                                    ft.ControlState.PRESSED: "#4990e2",
                                    ft.ControlState.DISABLED: "#cccccc"
                                }
                        ))
                        ],
                    bgcolor= '#EB2939'
                )
                page.open(alerta)
            else:
                page.snack_bar = ft.SnackBar(
                    ft.Text(message, color='#FFFFFF',weight='bold'),
                    bgcolor=color,
                    open=True,
                )



        def esconder_avisos():
            if page.snack_bar:
                page.snack_bar.open = False
                page.update()


        def arquivo(e: ft.FilePickerResultEvent, diferentes):
            if e.path:
                try:
                    if not e.path.endswith(".xlsx"):
                        e.path += ".xlsx"

                    wb = openpyxl.Workbook()
                    ws = wb.active

                    ws['A1'] = 'BOP'
                    ws['B1'] = 'Data Fato'
                    ws['C1'] = 'Municipio'
                    ws['D1'] = 'Consolidado'
                    ws['E1'] = 'Relato SISP'
                    ws['F1'] = 'Relato SiCADFULL'
                    ws['G1'] = 'Unidade Origem'
                    ws['H1'] = 'Unidade Responsavel'
                    ws['I1'] = 'Verificacao'



                    row_num = 2
                    for diff in set(diferentes):
                        ws[f'A{row_num}'] = diff[0]
                        ws[f'B{row_num}'] = diff[1]
                        ws[f'C{row_num}'] = diff[8]
                        ws[f'D{row_num}'] = diff[4]
                        ws[f'E{row_num}'] = diff[2]
                        ws[f'F{row_num}'] = diff[3]
                        ws[f'G{row_num}'] = diff[6]
                        ws[f'H{row_num}'] = diff[5]
                        ws[f'I{row_num}'] = diff[7]
                        row_num += 1

                    wb.save(e.path)
                    esconder_avisos()  
                    carregando.visible = False  
                    aviso("Arquivo salvo com sucesso!", color="green")
                    page.update()
                except Exception as ex:
                    esconder_avisos() 
                    carregando.visible = False 
                    page.update()
                    aviso(f"Erro ao salvar o arquivo: {ex}", color="red")




        def salvar_arquivo(e: ft.FilePickerResultEvent):
            arquivo(e, salvar_arquivo.diferentes)



        def atualizar_sugestoes(e, consolidado_entrada, page):
            sugestoes.visible = True
            opcoes = [opcao for opcao in opcoes_cosolidado if consolidado_entrada.value.lower() in opcao.lower()]
            sugestoes.controls.clear()
            for opcao in opcoes:
                bloco_lista = ft.ListTile(
                    title=ft.Text(opcao),
                    text_color="#000000",
                    hover_color="#E4E4E4",
                    bgcolor_activated='#0064FF'
                    )
                # Corrigir o comportamento do on_click para capturar a opção correta
                bloco_lista.on_click = lambda e, opt=opcao: opcao_selecionada(e, consolidado_entrada, page, opt)
                sugestoes.controls.append(bloco_lista)
            page.update()



        def opcao_selecionada(e, consolidado_entrada, page, opcao):
    #Atualizar o campo de input com a opção selecionada
            consolidado_entrada.value = opcao
            sugestoes.visible = False  
            page.update()


        def fechar(e, page):
            sugestoes.visible = False
            page.update()


        def pesquisando(e):
            aviso("Pesquisando...", color="blue")
            carregando.visible = True  
            page.update()

            if ano_entrada.value == '' and mes_menu not in meses:
                fechar(e,page)
                esconder_avisos() 
                carregando.visible = False  
                aviso("Por favor, insira um ano válido e selecione um mês.", color="red")
                page.update()
                return
            elif ano_entrada.value == '' or not ano_entrada.value.isdigit() or len(ano_entrada.value) > 4 or len(ano_entrada.value) < 4:
                fechar(e,page)
                esconder_avisos() 
                carregando.visible = False  
                aviso("Por favor, insira um ano válido.", color="red")
                page.update()
                return
            elif mes_menu.value not in meses:
                fechar(e,page)
                esconder_avisos() 
                carregando.visible = False  
                aviso("Por favor, selecione um mês.", color="red")
                page.update()
                return
            elif consolidado_entrada.value == '' or consolidado_entrada.value not in lista_consolidado:
                fechar(e,page)
                esconder_avisos() 
                carregando.visible = False  
                aviso("Por favor, insira o consolidado válido.", color="red")
                page.update()
                return
            
            
            ano = ano_entrada.value.strip()
            mes = mes_menu.value.strip().upper()
            consolidado = consolidado_entrada.value

            print('='*300)
            print(ano, mes, consolidado)
            print()

            try:
                pg_conn = psycopg2.connect(
            host=PG_HOST,
            database=PG_DBNAME,
            user=PG_USER,
            password=PG_PASSWORD
            )
                pg_cur = pg_conn.cursor()

                if consolidado == 'DESAPARECIMENTO DE PESSOA':
                    pg_cur.execute(sicad_desaparecimento,(consolidado,ano,mes))
                else:
                    pg_cur.execute(sicad,(consolidado,ano,mes))

                pg_data = []

                for row in pg_cur.fetchall():
                    bop = row[0]
                    data_fato = row[1]
                    relato = re.sub(r'<.*?>','', unidecode(row[2]))
                    relato2 = relato.replace('&NBSP;','')
                    relato2 = relato2.split()
                    relato2 = ' '.join(relato2)
                    consolidado = row[3]
                    unidade_responsavel = row[4]
                    unidade_origem = row[5]
                    municipio = row[6]
                    pg_data.append((bop, data_fato, relato2 ,consolidado, unidade_responsavel, unidade_origem,municipio))

                print()
                print('SICAD')
                print(pg_data)
                print()

                # pg_cur.close()
                pg_conn.close()

                bop_values = tuple(bop for bop,*_ in pg_data)
                bop_values_str = ", ".join([f"'{bop}'" for bop in bop_values])

                ora_conn = sisp_
                ora_cur = ora_conn.cursor()
                ora_cur.execute(sisp % bop_values_str)

                ora_data = []
                
                for o_row in ora_cur.fetchall():
                    bop = o_row[0]
                    data_fato = o_row[1]
                    relato_clob = re.sub(r'<.*?>','',unidecode(o_row[2].read().upper()))
                    relato_t = relato_clob.replace('&NBSP;','')
                    relato_t = relato_t.split()
                    relato_t = ' '.join(relato_t)
                    ora_data.append((bop, data_fato, relato_t))

                # ora_cur.close()
                # ora_conn.close()
                print()
                print('SISP')
                print(ora_data)
                print()
            
                diferentes = []

                for pg_row in pg_data:
                    for ora_row in ora_data:
                        if pg_row[0] == ora_row[0] and pg_row[2] != ora_row[2]:
                            char_diff = abs(len(pg_row[2]) - len(ora_row[2]))

                            if char_diff >= 90:
                                observacao = "Possível aditamento"
                            elif char_diff <= 90:
                                observacao = "Não tenho certeza"

                            diferentes.append((ora_row[0], pg_row[1], ora_row[2], pg_row[2], pg_row[3], pg_row[5],pg_row[4], observacao, pg_row[6]))

                if not diferentes:
                    esconder_avisos()  
                    carregando.visible = False  
                    aviso("Nenhuma diferença encontrada.", color="orange")
                    page.update()
                else:
                    salvar_arquivo.diferentes = diferentes
                    file_picker.save_file(
                        dialog_title="Salvar arquivo Excel",
                        file_type="xlsx",
                        file_name=f"Diferencas_{consolidado}_{ano}_{mes}.xlsx"
                    )
            except psycopg2.Error as pg_err:
                esconder_avisos()  
                carregando.visible = False  
                page.update()
                aviso(f"Erro no PostgreSQL: {pg_err}", color="red")
            except cx_Oracle.DatabaseError as ora_err:
                esconder_avisos()  
                carregando.visible = False  
                page.update()
                aviso(f"Erro no Oracle: {ora_err}", color="red")
            except Exception as ex:
                esconder_avisos()  
                carregando.visible = False  
                page.update()
                aviso(f"Erro inesperado: {ex}", color="red")





        file_picker.on_result = salvar_arquivo

        img = ft.Image(
        src=r"apps\assets\icone.svg",
        width= 80,
        height= 80
        )


        ano_entrada = ft.TextField(
        label="Escolha o ano", 
        width=169,
        color="#000000",
        border_color="#CED4DA",
        border_width=1,
        border_radius=2,
        label_style=ft.TextStyle(
            color="#000000",
            weight="w600",
            font_family="inter"
        ),
        text_style=ft.TextStyle(
            weight="w500",
            font_family="inter"
        ),
        on_focus= lambda e: fechar(e, page),
        focused_border_color='#0064FF'
        )


        mes_menu = ft.Dropdown(
        label="Mês",
        options=[
            ft.dropdown.Option(mes) for mes in meses.keys()
            ],
        width=166,
        bgcolor="#FFFFFF",
        color="#000000",
        border_color="#CED4DA",
        border_width=1,
        border_radius=2,
        label_style=ft.TextStyle(
            color="#000000",
            weight="w600",
            font_family="inter"
        ),
        text_style=ft.TextStyle(
            weight="w500",
            font_family="inter",
        ),
        on_focus= lambda e: fechar(e, page),
        focused_border_color='#0064FF'
        )





        consolidado_entrada = ft.TextField(
        label="Consolidado",
        on_change=lambda e: atualizar_sugestoes(e, consolidado_entrada, page),
        width=345,
        color="#000000",
        border_color="#CED4DA",
        border_width=1,
        border_radius=2,
        label_style=ft.TextStyle(
            color="#000000",
            weight="w600",
            font_family="inter",
        ),
        text_style=ft.TextStyle(
            weight="w500",
            font_family="inter",
        ),
        focused_border_color='#0064FF'
        )




        opcoes_cosolidado = lista_consolidado





        estilo_bt=ft.ButtonStyle(
        shape=ft.RoundedRectangleBorder(radius=2),
        bgcolor={
            ft.ControlState.DEFAULT: "#0064FF",
            ft.ControlState.FOCUSED: "#5caee0",
            ft.ControlState.PRESSED: "#4990e2",
            ft.ControlState.DISABLED: "#cccccc"
        },
        color={
            ft.ControlState.DEFAULT: ft.colors.WHITE,
            ft.ControlState.PRESSED: ft.colors.WHITE60,
        },
        elevation={
            ft.ControlState.DEFAULT: 3,
            ft.ControlState.HOVERED: 5,
        },
        padding=15,
        text_style={
            ft.ControlState.DEFAULT: ft.TextStyle(
                size=15,
                weight="w500"
            )
        },
        
        )



        bt_pesquisar = ft.ElevatedButton(
        "Pesquisar",
        icon="search",
        width=345,
        style=estilo_bt,
        on_click=pesquisando)


        carregando = ft.ProgressRing(
        visible=False,
        color='#0064FF'
        )


        app_bar= ft.TextButton(content=ft.Row([ft.Icon(name=ft.icons.ARROW_BACK_IOS_NEW, color='#0464f6'),ft.Text('Voltar',color='#0464f6')]), width=100, on_click=lambda e: page.go("/home"))
        i = ft.TextButton(content=ft.Row([ft.Icon(name=ft.icons.INFO_OUTLINE, color='#0464f6')]), on_click=lambda e: informacao_app(e))

        def informacao_app(e):
                dialogo = ft.AlertDialog(
                title=ft.Text("Informação"),
                content=ft.Container(
                    padding=5,
                    height=150,
                    width=300,
                    bgcolor='#e1e3eb',
                    border_radius=10,
                    content=ft.Column(
                        [
                            ft.Text(
                                'O aplicativo serve para comparar os relatos que estão na base local com os que estão na base do SISP D-1. Caso haja diferença, ele retorna um arquivo Excel contendo ambos os relatos.',
                                weight='w500'
                                )
                                ],
                                ft.MainAxisAlignment.CENTER,
                    )
                ), 
                adaptive=True,
                actions=[ft.TextButton("Fechar", on_click=lambda e: page.close(dialogo))],
                actions_alignment=ft.MainAxisAlignment.END,

            )
                page.dialog = dialogo
                dialogo.open = True
                page.update()


        body = ft.Container(
            expand=True,
            alignment=ft.alignment.center,
            content=ft.Column(
                alignment=ft.MainAxisAlignment.CENTER,
                controls=[
                    ft.Row([img],ft.MainAxisAlignment.CENTER),
                    # ft.Divider(height=20, color="transparent"),
                    ft.Row([ano_entrada,mes_menu],ft.MainAxisAlignment.CENTER),
                    ft.Row([consolidado_entrada],ft.MainAxisAlignment.CENTER),
                    ft.Row([sugestoes],ft.MainAxisAlignment.CENTER),
                    ft.Row([bt_pesquisar],ft.MainAxisAlignment.CENTER),
                    ft.Row([carregando],ft.MainAxisAlignment.CENTER),
                ]
            )
        )


        self.controls = [
            ft.SafeArea(
                expand=True,
                content=ft.Column(
                    controls=[ft.Row([app_bar, i],ft.MainAxisAlignment.SPACE_BETWEEN),body],spacing=0
                )
            )
        ]